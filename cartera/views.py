from django.shortcuts import render
import pandas as pd
import requests
from django.http import HttpResponse
import json
import ast
from openpyxl import Workbook
from .forms import MiFormulario

# Create your views here.
def cartera_show(request):
    form = MiFormulario()
    return render(request, 'cartera/index.html', {'form': form})

def cartera_vencida(df):
    aux = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for index, row in df.iterrows():
        if row[2] != None:
                dif = row[1] - row[2]
                if int(dif.days) > 365:
                    aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1
    return aux

def vigente(df):
    aux = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for index, row in df.iterrows():
        if row[2] != None:
            if row[2].year == row[1].year:
                if row[2].month == row[1].month:
                    aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1
    return aux

def adelantado(df):
    aux = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for index, row in df.iterrows():
        if row[2] != None:
            dif = row[1] - row[2]
            if row[1] < row[2]:
                if row[2].year == row[1].year:
                    if row[2].month != row[1].month:
                        aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1
                else:
                    if row[2].month != row[1].month or int(dif.days) < -30:
                        aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1                  
    return aux

def recuperado(df):
    aux = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for index, row in df.iterrows():
        if row[2] != None:
            dif = row[1] - row[2]
            if row[2].year == row[1].year:
                if row[2].month < row[1].month:
                    aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1
            else:
                if int(dif.days) > 0 and int(dif.days) < 365 :
                    aux[int(row[1].month) - 1] = aux[int(row[1].month) - 1] + 1
    return aux

def analisis(df):
    df['FPago'] = pd.to_datetime(df['FPago'])
    df['FFCAnt'] = pd.to_datetime(df['FFCAnt'])
    mes = ['Ene','Feb','Mar','Abr' , 'May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    su = pd.DataFrame(list(zip(mes, adelantado(df), vigente(df), recuperado(df), cartera_vencida(df))), columns = ['Mes','Adelantado', 'Vigente' ,'Recuperado', 'CarV'])
    su.loc[len(su)] = ['Total' , sum(su['Adelantado']), sum(su['Vigente']), sum(su['Recuperado']), sum(su['CarV'])]
    return su.to_dict('records')

def calculo(request):
    if request.POST["year"].isdigit() and  request.method == 'POST':
        year = int(request.POST ["year"])
        branch = request.POST.get('combo_box')
        url = "https://www.sistemapmn.com//api-app/internal/get-renews-by-year-and-branch"
        payload = json.dumps({
            "branch": branch,
            "year": year
        })
        headers = {
            'Content-Type': 'application/json',
            'Cookie': 'XSRF-TOKEN=eyJpdiI6IkpjNWNhRzhNdGRDaFgzMGdJNm9JOEE9PSIsInZhbHVlIjoicnlOUTdZQ3l2Z0dKSkUrbzFkTXNEZEpmdzlzMVhcL2V3dlA5azRQaGFxbjhYOTJJVCtWaUppaGVRWTNTNmZEU3I5UVBnRU5oXC9zK0ZhWXZRalFvamJXZz09IiwibWFjIjoiODUwZmQ1YjI4Njc1ZWI1ZDNhZDc1MjFmNTVkNmI5N2I4Njc1MDA4NDdiZGVhNWFhNmE3Mjg1M2JmMTJmMmM1YSJ9; laravel_session=eyJpdiI6IlZYSFlySUZ4aFFlQXFXTzAwS1JSVkE9PSIsInZhbHVlIjoiVFFBT0ZEUVJINml6MFZDOXNxVEJkT2lMM09aaUU2Y1paYlRPcHZ0WDExYjhDb0Q1WmRSbDNrQjJuUjZNK3hSU3BkcDBYdGdYWmdXMlhJSkJBZGpvXC93PT0iLCJtYWMiOiJiZjk5MDg0ODNlNTUwMmVlMGRjN2NlNGIwNjI3ZTQyNmQzYzlhYmE5YzY5ZmZlNDkxYTQ0OTFiNWNkYTNiMDBjIn0%3D'
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        info = analisis(pd.read_json(response.text))
        return render(request, 'cartera/analisis.html', {'info': info, 'year': year, 'branch': branch})
    mensaje = "Te falto llenar o llenaste incorrectamente"
    return HttpResponse(mensaje)

def download_excel(request):
    variable = request.GET.get('info')
    year = request.GET.get('year')
    branch = request.GET.get('branch')
    diccionario = ast.literal_eval(variable.replace(' ', ''))
    keysDic = ['Mes', 'Adelantado', 'Vigente', 'Recuperado', 'CarV']
    colE = ['A', 'B', 'C', 'D', 'E']
    cont = 2
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Mes'
    ws['B1'] = 'Adelantado'
    ws['C1'] = 'Vigente'
    ws['D1'] = 'Recuperado'
    ws['E1'] = 'Cartera Vencida'
    
    for i in range(13):
        for j in range(5):
            ws[colE[j] + str(cont)] = diccionario[i][keysDic[j]]
        cont = cont + 1

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="' + year+branch + '.xlsx"'

    wb.save(response)
    return response